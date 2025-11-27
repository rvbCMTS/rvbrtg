import logging
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

from Rapportering.Remittentstod.constants import (
    REPORT_TEMPLATE_PATH_PER_MODALITY,
    MODALITY_CT,
    MODALITY_DX,
    MODALITY_MG,
    MODALITY_XA,
    REPORT_OUTPUT_DIR, EXAM_GROUPING_RULES_BY_MODALITY,
    EXAM_GROUPING_TYPE_PROCEDURE_CODE, EXAM_GROUPING_TYPE_PROTOCOL_CODE,
    OUTPUT_COL_EXAM, VALID_SERIES_COLUMNS, VALID_STUDY_COLUMNS,
    MG_COL_PROJECTION, MG_COL_EXAM_INDEX, MG_COL_EXAM_TYPE, OUTPUT_COL_WEIGTH_CATEGORY,
    OUTPUT_COL_EFFECTIVE_DOSE, OUTPUT_COL_BODY_PART)

logger = logging.getLogger("yearly_statistics")


def save_formatted_data(data: pd.DataFrame, modality: str) -> None:
    REPORT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    logger.info(f"Sparar rapporter i {REPORT_OUTPUT_DIR.absolute()}")

    for exam_name in data[OUTPUT_COL_EXAM].unique().tolist():
        logger.debug(f"Skapar rapport för undersökning {exam_name}")
        _create_report_main(data=data, exam_name=exam_name, modality=modality)

    return


def _create_report_main(data: pd.DataFrame, modality:str, exam_name:str):
    tmp_data = data[(data[OUTPUT_COL_EXAM] == exam_name)].reset_index()

    if modality not in [MODALITY_DX]:
        raise NotImplementedError(f"Modality '{modality}' not implemented.")
    
    output_path: Path = REPORT_OUTPUT_DIR / f"{modality} - {exam_name}.pdf"
    if output_path.exists():
        try:
            output_path.unlink()
        except Exception:
            logger.error(f"Kunde inte ta bort befintlig rapportfil: {output_path}")
            
    with PdfPages(output_path) as pdf:
        try:
            if modality == MODALITY_DX:
                _create_report_dx(pdf=pdf, data=tmp_data, exam_name=exam_name)
        except Exception:
            logger.error(f"Kunde inte skapa rapport för {modality} - {exam_name}")

def _create_report_ct(report_sheet, data: pd.DataFrame):
    # Sort data based on their absolute diff in CDTI compared to the median of all exams in order to report the middle
    # interval of all exams
    data['CTDIdiff'] = (
            data[VALID_SERIES_COLUMNS.MeanCTDIvol] - data.loc[:, VALID_SERIES_COLUMNS.MeanCTDIvol].median()
    ).abs()
    data.sort_values('CTDIdiff', inplace=True, ignore_index=True)

    for row_ind in range(20):
        report_sheet.cell(row=row_ind + 3, column=1).value = data[VALID_SERIES_COLUMNS.MeanCTDIvol][row_ind]
        report_sheet.cell(row=row_ind + 3, column=2).value = data[VALID_SERIES_COLUMNS.DlPv][row_ind]
        report_sheet.cell(row=row_ind + 3, column=3).value = data[VALID_SERIES_COLUMNS.SizeSpecificDoseEstimation][row_ind]
        report_sheet.cell(row=row_ind + 3, column=4).value = data[VALID_STUDY_COLUMNS.PatientAge][row_ind]
        report_sheet.cell(row=row_ind + 3, column=5).value = data[VALID_STUDY_COLUMNS.PatientsSex][row_ind]
        report_sheet.cell(row=row_ind + 3, column=6).value = data[VALID_STUDY_COLUMNS.PatientsSize][row_ind]
        report_sheet.cell(row=row_ind + 3, column=7).value = data[VALID_STUDY_COLUMNS.PatientsWeight][row_ind]
        report_sheet.cell(row=row_ind + 3, column=8).value = data[VALID_STUDY_COLUMNS.Machine][row_ind]
        report_sheet.cell(row=row_ind + 3, column=9).value = data[VALID_SERIES_COLUMNS.AcquisitionProtocol][row_ind]

    return report_sheet


def _create_report_dx(pdf, data: pd.DataFrame, exam_name: str):
    # Plot a histogram of effective dose and calculate basic statistics
    agg_data = data.groupby(by=OUTPUT_COL_EXAM).agg(
                Antal=pd.NamedAgg(column=OUTPUT_COL_EFFECTIVE_DOSE, aggfunc="count"),
                Dose_mean=pd.NamedAgg(column=OUTPUT_COL_EFFECTIVE_DOSE, aggfunc="mean"),
                Dose_median=pd.NamedAgg(column=OUTPUT_COL_EFFECTIVE_DOSE, aggfunc="median"),
                Dose_95=pd.NamedAgg(column=OUTPUT_COL_EFFECTIVE_DOSE, aggfunc=lambda x: x.quantile(0.95))
    )

    fig, ax = plt.subplots(1, 1, sharey=True, tight_layout=True)
    ax.hist(data[OUTPUT_COL_EFFECTIVE_DOSE], bins=round(len(data) ** (1/2)))

    plt.text(data[OUTPUT_COL_EFFECTIVE_DOSE].max()*0.4,
             ax.get_yticks()[1],
             f"Body part (DAP -> mSv):\n {data[OUTPUT_COL_BODY_PART].unique().tolist()}\n" +
             f"Study description:\n {data[VALID_STUDY_COLUMNS.StudyDescription].unique().tolist()}\n" +
             f"Machine:\n {_list_to_multiline_string(data[VALID_STUDY_COLUMNS.Machine].unique().tolist())}\n\n" +
             f"Antal undersökningar: {agg_data['Antal'].values[0]}\n" +
             f"Medelvärde: {agg_data['Dose_mean'].values[0]:.2f} mSv\n" +
             f"Median: {agg_data['Dose_median'].values[0]:.2f} mSv\n" +
             f"95-percentil: {agg_data['Dose_95'].values[0]:.2f} mSv"
    )

    ax.set_ylabel("Antal undersökningar")
    ax.set_xlabel("Effektiv dos (mSv)")
    ax.set_title(f"Histogram över effektiv dos för {exam_name}")

    plt.show()
    pdf.savefig(fig)
    plt.close(fig)


def _create_report_xa(report_sheet, data: pd.DataFrame):
    # Sort data based on their absolute diff in DAP compared to the median of all exams in order to report the middle
    # interval of all exams
    data['DAPdiff'] = (
            data[VALID_STUDY_COLUMNS.DoseAreaProductTotal] - data.loc[:, VALID_STUDY_COLUMNS.DoseAreaProductTotal].median()
    ).abs()
    data.sort_values('DAPdiff', inplace=True, ignore_index=True)

    for row_ind in range(0, min(20, len(data))):
        report_sheet.cell(row=row_ind + 3, column=1).value = data[VALID_STUDY_COLUMNS.FluoroDoseAreaProductTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=2).value = data[VALID_STUDY_COLUMNS.FluoroDoseRPTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=3).value = data[VALID_STUDY_COLUMNS.TotalFluoroTime][row_ind]
        report_sheet.cell(row=row_ind + 3, column=4).value = data[VALID_STUDY_COLUMNS.AcquisitionDoseAreaProductTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=5).value = data[VALID_STUDY_COLUMNS.AcquisitionDoseRPTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=6).value = data[VALID_STUDY_COLUMNS.TotalNumberOfIrradiationEvents][row_ind]
        report_sheet.cell(row=row_ind + 3, column=7).value = data[VALID_STUDY_COLUMNS.PatientAge][row_ind]
        report_sheet.cell(row=row_ind + 3, column=8).value = data[VALID_STUDY_COLUMNS.PatientsSex][row_ind]
        report_sheet.cell(row=row_ind + 3, column=9).value = data[VALID_STUDY_COLUMNS.PatientsSize][row_ind]
        report_sheet.cell(row=row_ind + 3, column=10).value = data[VALID_STUDY_COLUMNS.PatientsWeight][row_ind]

    return report_sheet


def _create_report_mg(report_sheet, data: pd.DataFrame, machine: str, exam_name: str):
    # Sort data based on their absolute diff in AGD compared to the mean of all exams in order to report the middle
    # interval of all exams
    COL_AGD_SUM = "AGD_SUM"
    data[COL_AGD_SUM] = data.groupby(VALID_STUDY_COLUMNS.Id)[VALID_SERIES_COLUMNS.AverageGlandularDose].transform("sum")
    data['AGDdiff'] = (
            data[COL_AGD_SUM] - data.groupby(VALID_STUDY_COLUMNS.Id)[COL_AGD_SUM].max().median()
    ).abs()
    exam_ids = data.sort_values(['AGDdiff'], inplace=False, ignore_index=True)[VALID_STUDY_COLUMNS.Id].unique().tolist()
    main_row_ind = 3
    if len(exam_ids) < 20:
        logger.warning(f"Too few exams in category {exam_name} ({machine})")
        return report_sheet
    for exam_ind in range(20):
        tmp_data = data[data[VALID_STUDY_COLUMNS.Id] == exam_ids[exam_ind]].sort_values([MG_COL_EXAM_INDEX], ignore_index=True)
        for row_ind in range(len(tmp_data)):
            report_sheet.cell(row=main_row_ind, column=1).value = exam_ind+1
            report_sheet.cell(row=main_row_ind, column=2).value = tmp_data[VALID_SERIES_COLUMNS.AverageGlandularDose][row_ind]
            report_sheet.cell(row=main_row_ind, column=3).value = tmp_data[MG_COL_PROJECTION][row_ind]
            report_sheet.cell(row=main_row_ind, column=4).value = tmp_data[VALID_SERIES_COLUMNS.kVp][row_ind]
            report_sheet.cell(row=main_row_ind, column=5).value = tmp_data[VALID_SERIES_COLUMNS.Exposure][row_ind]
            report_sheet.cell(row=main_row_ind, column=6).value = tmp_data[VALID_SERIES_COLUMNS.CompressionForce][row_ind]
            report_sheet.cell(row=main_row_ind, column=7).value = tmp_data[VALID_SERIES_COLUMNS.CompressionThickness][row_ind]
            report_sheet.cell(row=main_row_ind, column=8).value = tmp_data[VALID_SERIES_COLUMNS.AnodeTargetMaterial][row_ind]
            report_sheet.cell(row=main_row_ind, column=9).value = tmp_data[VALID_SERIES_COLUMNS.XrayFilterMaterial][row_ind]

            #report_sheet.cell(row=main_row_ind, column=1).value = tmp_data[MG_COL_EXAM_INDEX][row_ind]
            #report_sheet.cell(row=main_row_ind, column=2).value = tmp_data[MG_COL_PROJECTION][row_ind]
            #report_sheet.cell(row=main_row_ind, column=3).value = tmp_data[VALID_SERIES_COLUMNS.kVp][row_ind]
            #report_sheet.cell(row=main_row_ind, column=4).value = tmp_data[VALID_SERIES_COLUMNS.Exposure][row_ind]
            #report_sheet.cell(row=main_row_ind, column=5).value = tmp_data[VALID_SERIES_COLUMNS.CompressionForce][row_ind]
            #report_sheet.cell(row=main_row_ind, column=6).value = tmp_data[VALID_SERIES_COLUMNS.AverageGlandularDose][row_ind]
            #report_sheet.cell(row=main_row_ind, column=7).value = tmp_data[VALID_SERIES_COLUMNS.CompressionThickness][row_ind]
            #report_sheet.cell(row=main_row_ind, column=8).value = tmp_data[VALID_SERIES_COLUMNS.AnodeTargetMaterial][row_ind]
            #report_sheet.cell(row=main_row_ind, column=9).value = tmp_data[VALID_SERIES_COLUMNS.XrayFilterMaterial][row_ind]

            main_row_ind = main_row_ind + 1

    return report_sheet

def _list_to_multiline_string(lst, per_line=5, sep=", "):
    chunks = [sep.join(map(str, lst[i:i+per_line])) for i in range(0, len(lst), per_line)]
    return "\n".join(chunks)