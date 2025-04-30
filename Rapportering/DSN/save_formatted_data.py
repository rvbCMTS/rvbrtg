import logging
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from Rapportering.DSN.constants import (
    REPORT_TEMPLATE_PATH_PER_MODALITY,
    MODALITY_CT,
    MODALITY_DX,
    MODALITY_MG,
    MODALITY_XA,
    REPORT_OUTPUT_DIR, EXAM_GROUPING_RULES_BY_MODALITY,
    EXAM_GROUPING_TYPE_PROCEDURE_CODE, EXAM_GROUPING_TYPE_PROTOCOL_CODE,
    OUTPUT_COL_EXAM, VALID_SERIES_COLUMNS, VALID_STUDY_COLUMNS,
    MG_COL_PROJECTION, MG_COL_EXAM_INDEX, MG_COL_EXAM_TYPE, OUTPUT_COL_WEIGTH_CATEGORY)

logger = logging.getLogger("yearly_statistics")


def save_formatted_data(data: pd.DataFrame, modality: str) -> None:
    report_template = REPORT_TEMPLATE_PATH_PER_MODALITY[modality]
    REPORT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    logger.info(f"Sparar rapporter i {REPORT_OUTPUT_DIR.absolute()}")

    for exam_name in data[OUTPUT_COL_EXAM].unique().tolist():
        for machine in data[data[OUTPUT_COL_EXAM] == exam_name][VALID_STUDY_COLUMNS.Machine].unique().tolist():
            if modality == MODALITY_MG:
                logger.debug(f"Skapar rapport för {modality} ({machine}) kopplade till undersökning {exam_name}")
                _create_report_main(template_path=report_template, data=data, machine=machine, exam_name=exam_name, weight=None, modality=modality)
            else:
                for weight in data[(data[OUTPUT_COL_EXAM] == exam_name) & (data[VALID_STUDY_COLUMNS.Machine] == machine)][OUTPUT_COL_WEIGTH_CATEGORY].unique().tolist():
                    logger.debug(f"Skapar rapport för {modality} ({machine}) kopplade till undersökning {exam_name}")

                    _create_report_main(template_path=report_template, data=data, machine=machine, exam_name=exam_name, weight=weight, modality=modality)

    return


def _create_report_main(template_path: Path, data: pd.DataFrame, modality:str, machine: str, exam_name:str, weight:str):
    output_path: Path = REPORT_OUTPUT_DIR / f"{modality} - {machine} - {exam_name} - {weight}{template_path.suffix}"

    report_template = load_workbook(template_path)
    sheet = report_template.active

    tmp_data = data[(data[OUTPUT_COL_EXAM] == exam_name) & (data[VALID_STUDY_COLUMNS.Machine] == machine)].reset_index()
    if modality != MODALITY_MG:
        tmp_data = tmp_data[data[OUTPUT_COL_WEIGTH_CATEGORY] == weight].reset_index()

    if modality not in [
        MODALITY_DX, MODALITY_MG, MODALITY_XA, MODALITY_CT
    ]:
        raise NotImplementedError(f"Modality '{modality}' not implemented.")

    sheet = (
        _create_report_ct(report_sheet=sheet, data=tmp_data)
        if modality == MODALITY_CT else (
            _create_report_dx(report_sheet=sheet, data=tmp_data)
            if modality == MODALITY_DX else (
                _create_report_mg(report_sheet=sheet, data=tmp_data, machine=machine, exam_name=exam_name)
                if modality == MODALITY_MG else (
                    _create_report_xa(report_sheet=sheet, data=tmp_data)
                )
            )
        )
    )
    report_template.save(output_path)
    report_template.close()


def _create_report_ct(report_sheet, data: pd.DataFrame):
    # Sort data based on their absolute diff in CDTI compared to the median of all exams in order to report the middle
    # interval of all exams
    data['CTDIdiff'] = (
            data[VALID_SERIES_COLUMNS.MeanCTDIvol] - data.loc[:, VALID_SERIES_COLUMNS.MeanCTDIvol].median()
    ).abs()
    data.sort_values('CTDIdiff', inplace=True, ignore_index=True)

    for row_ind in range(20):
        report_sheet.cell(row=row_ind + 3, column=1).value = data[VALID_SERIES_COLUMNS.MeanCTDIvol][row_ind]
        report_sheet.cell(row=row_ind + 3, column=2).value = data[VALID_SERIES_COLUMNS.DlPv][row_ind]
        report_sheet.cell(row=row_ind + 3, column=3).value = data[VALID_SERIES_COLUMNS.SizeSpecificDoseEstimation][row_ind]
        report_sheet.cell(row=row_ind + 3, column=4).value = data[VALID_STUDY_COLUMNS.PatientAge][row_ind]
        report_sheet.cell(row=row_ind + 3, column=5).value = data[VALID_STUDY_COLUMNS.PatientsSex][row_ind]
        report_sheet.cell(row=row_ind + 3, column=6).value = data[VALID_STUDY_COLUMNS.PatientsSize][row_ind]
        report_sheet.cell(row=row_ind + 3, column=7).value = data[VALID_STUDY_COLUMNS.PatientsWeight][row_ind]
        report_sheet.cell(row=row_ind + 3, column=8).value = data[VALID_STUDY_COLUMNS.Machine][row_ind]
        report_sheet.cell(row=row_ind + 3, column=9).value = data[VALID_SERIES_COLUMNS.AcquisitionProtocol][row_ind]

    return report_sheet


def _create_report_dx(report_sheet, data: pd.DataFrame):
    # Sort data based on their absolute diff in DAP compared to the median of all exams in order to report the middle
    # interval of all exams
    data['DAPdiff'] = (
            data[VALID_STUDY_COLUMNS.DoseAreaProductTotal] - data.loc[:, VALID_STUDY_COLUMNS.DoseAreaProductTotal].median()
    ).abs()
    data.sort_values('DAPdiff', inplace=True, ignore_index=True)

    for row_ind in range(0, min(20, len(data))):
        report_sheet.cell(row=row_ind + 3, column=1).value = data[VALID_STUDY_COLUMNS.TotalNumberOfRadiographicFrames][row_ind]
        report_sheet.cell(row=row_ind + 3, column=2).value = data[VALID_STUDY_COLUMNS.DoseAreaProductTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=3).value = data[VALID_STUDY_COLUMNS.PatientAge][row_ind]
        report_sheet.cell(row=row_ind + 3, column=4).value = data[VALID_STUDY_COLUMNS.PatientsSex][row_ind]
        report_sheet.cell(row=row_ind + 3, column=5).value = data[VALID_STUDY_COLUMNS.PatientsSize][row_ind]
        report_sheet.cell(row=row_ind + 3, column=6).value = data[VALID_STUDY_COLUMNS.PatientsWeight][row_ind]

    return report_sheet


def _create_report_xa(report_sheet, data: pd.DataFrame):
    # Sort data based on their absolute diff in DAP compared to the median of all exams in order to report the middle
    # interval of all exams
    data['DAPdiff'] = (
            data[VALID_STUDY_COLUMNS.DoseAreaProductTotal] - data.loc[:, VALID_STUDY_COLUMNS.DoseAreaProductTotal].median()
    ).abs()
    data.sort_values('DAPdiff', inplace=True, ignore_index=True)

    for row_ind in range(0, min(20, len(data))):
        report_sheet.cell(row=row_ind + 3, column=1).value = data[VALID_STUDY_COLUMNS.FluoroDoseAreaProductTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=2).value = data[VALID_STUDY_COLUMNS.FluoroDoseRPTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=3).value = data[VALID_STUDY_COLUMNS.TotalFluoroTime][row_ind]
        report_sheet.cell(row=row_ind + 3, column=4).value = data[VALID_STUDY_COLUMNS.AcquisitionDoseAreaProductTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=5).value = data[VALID_STUDY_COLUMNS.AcquisitionDoseRPTotal][row_ind]
        report_sheet.cell(row=row_ind + 3, column=6).value = data[VALID_STUDY_COLUMNS.TotalNumberOfIrradiationEvents][row_ind]
        report_sheet.cell(row=row_ind + 3, column=7).value = data[VALID_STUDY_COLUMNS.PatientAge][row_ind]
        report_sheet.cell(row=row_ind + 3, column=8).value = data[VALID_STUDY_COLUMNS.PatientsSex][row_ind]
        report_sheet.cell(row=row_ind + 3, column=9).value = data[VALID_STUDY_COLUMNS.PatientsSize][row_ind]
        report_sheet.cell(row=row_ind + 3, column=10).value = data[VALID_STUDY_COLUMNS.PatientsWeight][row_ind]

    return report_sheet


def _create_report_mg(report_sheet, data: pd.DataFrame, machine: str, exam_name: str):
    # Sort data based on their absolute diff in AGD compared to the mean of all exams in order to report the middle
    # interval of all exams
    COL_AGD_SUM = "AGD_SUM"
    data[COL_AGD_SUM] = data.groupby(VALID_STUDY_COLUMNS.Id)[VALID_SERIES_COLUMNS.AverageGlandularDose].transform("sum")
    data['AGDdiff'] = (
            data[COL_AGD_SUM] - data.groupby(VALID_STUDY_COLUMNS.Id)[COL_AGD_SUM].max().median()
    ).abs()
    exam_ids = data.sort_values(['AGDdiff'], inplace=False, ignore_index=True)[VALID_STUDY_COLUMNS.Id].unique().tolist()
    main_row_ind = 3
    if len(exam_ids) < 20:
        logger.warning(f"Too few exams in category {exam_name} ({machine})")
        return report_sheet
    for exam_ind in range(20):
        tmp_data = data[data[VALID_STUDY_COLUMNS.Id] == exam_ids[exam_ind]].sort_values([MG_COL_EXAM_INDEX], ignore_index=True)
        for row_ind in range(len(tmp_data)):
            report_sheet.cell(row=main_row_ind, column=1).value = tmp_data[MG_COL_EXAM_INDEX][row_ind]
            report_sheet.cell(row=main_row_ind, column=2).value = tmp_data[MG_COL_PROJECTION][row_ind]
            report_sheet.cell(row=main_row_ind, column=3).value = tmp_data[VALID_SERIES_COLUMNS.kVp][row_ind]
            report_sheet.cell(row=main_row_ind, column=4).value = tmp_data[VALID_SERIES_COLUMNS.Exposure][row_ind]
            report_sheet.cell(row=main_row_ind, column=5).value = tmp_data[VALID_SERIES_COLUMNS.CompressionForce][row_ind]
            report_sheet.cell(row=main_row_ind, column=6).value = tmp_data[VALID_SERIES_COLUMNS.AverageGlandularDose][row_ind]
            report_sheet.cell(row=main_row_ind, column=7).value = tmp_data[VALID_SERIES_COLUMNS.CompressionThickness][row_ind]
            report_sheet.cell(row=main_row_ind, column=8).value = tmp_data[VALID_SERIES_COLUMNS.AnodeTargetMaterial][row_ind]
            report_sheet.cell(row=main_row_ind, column=9).value = tmp_data[VALID_SERIES_COLUMNS.XrayFilterMaterial][row_ind]

            main_row_ind = main_row_ind + 1

    return report_sheet
