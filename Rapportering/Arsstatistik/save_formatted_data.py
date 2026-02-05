import logging
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
import pandas as pd

from Rapportering.Arsstatistik.constants import (
    REPORT_TEMPLATE_PATH_PER_MODALITY,
    MODALITY_CT,
    REPORT_OUTPUT_DIR,
    MODALITY_DX,
    MODALITY_MG,
    MODALITY_XA,
    EXAM_GROUPING_RULES_BY_MODALITY,
    EXAM_GROUPING_TYPE_PROCEDURE_CODE,
    EXAM_GROUPING_TYPE_PROTOCOL_CODE,
    AGE_SEX_CATEGORY_JUNIOR_FEMALE,
    AGE_SEX_CATEGORY_JUNIOR_MALE,
    AGE_SEX_CATEGORY_ADULT_FEMALE_16_40, AGE_SEX_CATEGORY_ADULT_FEMALE_41_65, AGE_SEX_CATEGORY_ADULT_FEMALE_66plus,
    AGE_SEX_CATEGORY_DOSE_FEMALE, AGE_SEX_CATEGORY_ADULT_MALE_16_40, AGE_SEX_CATEGORY_ADULT_MALE_41_65,
    AGE_SEX_CATEGORY_ADULT_MALE_66plus, AGE_SEX_CATEGORY_DOSE_MALE, AGE_SEX_CATEGORY_DOSE_GIRL,
    AGE_SEX_CATEGORY_DOSE_BOY, EXAMS_EXEMPT_FROM_REPORTING_DOSE,
    OUTPUT_KEY_MEAN_DOSE,
    OUTPUT_KEY_MEDIAN_DOSE,
    OUTPUT_KEY_Q1_DOSE,
    OUTPUT_KEY_Q3_DOSE,
)

logger = logging.getLogger("yearly_statistics")


def save_formatted_data(data_count: pd.DataFrame, data_dose: pd.DataFrame, modality: str) -> None:
    report_template = REPORT_TEMPLATE_PATH_PER_MODALITY[modality]
    REPORT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    logger.info(f"Sparar rapporter i {REPORT_OUTPUT_DIR.absolute()}")

    for hospital in data_count.index.get_level_values(0).unique().tolist():
        logger.debug(f"Skapar rapport för {modality} kopplade till {hospital}")
        if modality == MODALITY_CT:
            _create_report_file_ct(template_path=report_template, data_count=data_count, data_dose=data_dose, hospital=hospital)
            continue
        if modality == MODALITY_DX:
            _create_report_file_dx(template_path=report_template, data_count=data_count, data_dose=data_dose, hospital=hospital)
            continue
        if modality == MODALITY_MG:
            _create_report_file_mg(template_path=report_template, data_count=data_count, data_dose=data_dose, hospital=hospital)
            continue
        if modality == MODALITY_XA:
            _create_report_file_xa(template_path=report_template, data_count=data_count, data_dose=data_dose, hospital=hospital)
            continue
        raise NotImplementedError(f"Report creation not implemented for {modality=}")

    return


def _get_exam_grouping_name_from_template_row(sheet, row: int) -> str | None:
    us_id = sheet.cell(row=row, column=1).value

    if not us_id:
        return None

    modality = sheet.cell(row=row, column=2).value
    aim = sheet.cell(row=row, column=3).value
    region = sheet.cell(row=row, column=4).value

    return f"{us_id}:{modality}:{aim}:{region}"


def _create_report_main(template_path: Path, data_count: pd.DataFrame, data_dose: pd.DataFrame, hospital: str, modality: str = "Inte mammo"):
    output_path: Path = REPORT_OUTPUT_DIR / f"{template_path.stem.split(' ')[0]} {hospital} DosReg{template_path.suffix}"

    report_template = load_workbook(template_path)
    sheet = report_template.active
    sheet_exams = { exam: row for row in range(3, 50) if (exam := _get_exam_grouping_name_from_template_row(sheet, row))}

    for exam, row in sheet_exams.items():
        if (hospital, exam) not in data_count.index:
            logger.warning(f"Ingen '{exam}' undersökning hittades kopplad till {hospital}")
            continue

        df_row = data_count.loc[(hospital, exam)]
        if not len(df_row):
            print(f"Hittade inte '{exam}' från template bland undersökningarna")
            continue

        df_row_dose = data_dose.loc[(hospital, exam)]

        if AGE_SEX_CATEGORY_ADULT_FEMALE_16_40 in df_row.index.levels[1]:
            sheet.cell(row=row, column=5).value = val if not np.isnan(
                (val := df_row["Antal"][AGE_SEX_CATEGORY_ADULT_FEMALE_16_40])) else "-"

        if AGE_SEX_CATEGORY_ADULT_FEMALE_41_65 in df_row.index.levels[1]:
            sheet.cell(row=row, column=6).value = val if not np.isnan(
                (val := df_row["Antal"][AGE_SEX_CATEGORY_ADULT_FEMALE_41_65])) else "-"

        if AGE_SEX_CATEGORY_ADULT_FEMALE_66plus in df_row.index.levels[1]:
            sheet.cell(row=row, column=7).value = val if not np.isnan(
                (val := df_row["Antal"][AGE_SEX_CATEGORY_ADULT_FEMALE_66plus])) else "-"

        if exam not in EXAMS_EXEMPT_FROM_REPORTING_DOSE and AGE_SEX_CATEGORY_DOSE_FEMALE in df_row_dose.index.levels[1]:
            sheet.cell(row=row, column=8).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_MEAN_DOSE][AGE_SEX_CATEGORY_DOSE_FEMALE])) else "-"
            sheet.cell(row=row, column=9).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_MEDIAN_DOSE][AGE_SEX_CATEGORY_DOSE_FEMALE])) else "-"
            sheet.cell(row=row, column=10).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_Q1_DOSE][AGE_SEX_CATEGORY_DOSE_FEMALE])) else "-"
            sheet.cell(row=row, column=11).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_Q3_DOSE][AGE_SEX_CATEGORY_DOSE_FEMALE])) else "-"

        if modality == MODALITY_MG:
            if AGE_SEX_CATEGORY_JUNIOR_FEMALE in df_row.index.levels[1]:
                sheet.cell(row=row, column=12).value = val if not np.isnan(
                    (val := df_row["Antal"][AGE_SEX_CATEGORY_JUNIOR_FEMALE])) else "-"

            if exam not in EXAMS_EXEMPT_FROM_REPORTING_DOSE and AGE_SEX_CATEGORY_DOSE_GIRL in df_row_dose.index.levels[
                1]:
                sheet.cell(row=row, column=13).value = val if not np.isnan(
                    (val := df_row_dose[OUTPUT_KEY_MEAN_DOSE][AGE_SEX_CATEGORY_DOSE_GIRL])) else "-"
                sheet.cell(row=row, column=14).value = val if not np.isnan(
                    (val := df_row_dose[OUTPUT_KEY_MEDIAN_DOSE][AGE_SEX_CATEGORY_DOSE_GIRL])) else "-"
                sheet.cell(row=row, column=15).value = val if not np.isnan(
                    (val := df_row_dose[OUTPUT_KEY_Q1_DOSE][AGE_SEX_CATEGORY_DOSE_GIRL])) else "-"
                sheet.cell(row=row, column=16).value = val if not np.isnan(
                    (val := df_row_dose[OUTPUT_KEY_Q3_DOSE][AGE_SEX_CATEGORY_DOSE_GIRL])) else "-"
            continue


        if AGE_SEX_CATEGORY_ADULT_MALE_16_40 in df_row.index.levels[1]:
            sheet.cell(row=row, column=12).value = val if not np.isnan(
                (val := df_row["Antal"][AGE_SEX_CATEGORY_ADULT_MALE_16_40])) else "-"

        if AGE_SEX_CATEGORY_ADULT_MALE_41_65 in df_row.index.levels[1]:
            sheet.cell(row=row, column=13).value = val if not np.isnan(
                (val := df_row["Antal"][AGE_SEX_CATEGORY_ADULT_MALE_41_65])) else "-"

        if AGE_SEX_CATEGORY_ADULT_MALE_66plus in df_row.index.levels[1]:
            sheet.cell(row=row, column=14).value = val if not np.isnan(
                (val := df_row["Antal"][AGE_SEX_CATEGORY_ADULT_MALE_66plus])) else "-"

        if exam not in EXAMS_EXEMPT_FROM_REPORTING_DOSE and AGE_SEX_CATEGORY_DOSE_MALE in df_row_dose.index.levels[1]:
            sheet.cell(row=row, column=15).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_MEAN_DOSE][AGE_SEX_CATEGORY_DOSE_MALE])) else "-"
            sheet.cell(row=row, column=16).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_MEDIAN_DOSE][AGE_SEX_CATEGORY_DOSE_MALE])) else "-"
            sheet.cell(row=row, column=17).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_Q1_DOSE][AGE_SEX_CATEGORY_DOSE_MALE])) else "-"
            sheet.cell(row=row, column=18).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_Q3_DOSE][AGE_SEX_CATEGORY_DOSE_MALE])) else "-"

        if AGE_SEX_CATEGORY_JUNIOR_FEMALE in df_row.index.levels[1]:
            sheet.cell(row=row, column=19).value = val if not np.isnan(
                (val := df_row["Antal"][AGE_SEX_CATEGORY_JUNIOR_FEMALE])) else "-"

        if exam not in EXAMS_EXEMPT_FROM_REPORTING_DOSE and AGE_SEX_CATEGORY_DOSE_GIRL in df_row_dose.index.levels[1]:
            sheet.cell(row=row, column=20).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_MEAN_DOSE][AGE_SEX_CATEGORY_DOSE_GIRL])) else "-"
            sheet.cell(row=row, column=21).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_MEDIAN_DOSE][AGE_SEX_CATEGORY_DOSE_GIRL])) else "-"
            sheet.cell(row=row, column=22).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_Q1_DOSE][AGE_SEX_CATEGORY_DOSE_GIRL])) else "-"
            sheet.cell(row=row, column=23).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_Q3_DOSE][AGE_SEX_CATEGORY_DOSE_GIRL])) else "-"

        if AGE_SEX_CATEGORY_JUNIOR_MALE in df_row.index.levels[1]:
            sheet.cell(row=row, column=24).value = val if not np.isnan(
                (val := df_row["Antal"][AGE_SEX_CATEGORY_JUNIOR_MALE])) else "-"

        if exam not in EXAMS_EXEMPT_FROM_REPORTING_DOSE and AGE_SEX_CATEGORY_DOSE_BOY in df_row_dose.index.levels[1]:
            sheet.cell(row=row, column=25).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_MEAN_DOSE][AGE_SEX_CATEGORY_DOSE_BOY])) else "-"
            sheet.cell(row=row, column=26).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_MEDIAN_DOSE][AGE_SEX_CATEGORY_DOSE_BOY])) else "-"
            sheet.cell(row=row, column=27).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_Q1_DOSE][AGE_SEX_CATEGORY_DOSE_BOY])) else "-"
            sheet.cell(row=row, column=28).value = val if not np.isnan(
                (val := df_row_dose[OUTPUT_KEY_Q3_DOSE][AGE_SEX_CATEGORY_DOSE_BOY])) else "-"

    report_template.save(output_path)
    report_template.close()


def _get_exam_codes_for_modality(modality: str) -> dict[str, list[str]]:
    procedure_codes = EXAM_GROUPING_RULES_BY_MODALITY[modality].get(EXAM_GROUPING_TYPE_PROCEDURE_CODE)
    protocol_codes = EXAM_GROUPING_RULES_BY_MODALITY[modality].get(EXAM_GROUPING_TYPE_PROTOCOL_CODE)

    if not procedure_codes and not protocol_codes:
        return {}

    if procedure_codes and protocol_codes:
        output = {key: val + protocol_codes[key] for key, val in procedure_codes.items() if key in list(protocol_codes.keys())}
        if any(only_protocol_code_keys := [key for key in list(protocol_codes.keys()) if key not in list(procedure_codes.keys())]):
            output = output | {key: protocol_codes[key] for key in only_protocol_code_keys}
        return output

    if procedure_codes:
        return procedure_codes

    return procedure_codes


def _create_report_file_ct(template_path: Path, data_count: pd.DataFrame, data_dose: pd.DataFrame, hospital: str) -> None:
    exam_codes = _get_exam_codes_for_modality(modality=MODALITY_CT)
    _create_report_main(template_path=template_path, data_count=data_count, data_dose=data_dose, hospital=hospital)


def _create_report_file_dx(template_path: Path, data_count: pd.DataFrame, data_dose: pd.DataFrame, hospital: str) -> None:
    exam_codes = _get_exam_codes_for_modality(modality=MODALITY_DX)
    _create_report_main(template_path=template_path, data_count=data_count, data_dose=data_dose, hospital=hospital)


def _create_report_file_mg(template_path: Path, data_count: pd.DataFrame, data_dose: pd.DataFrame, hospital: str) -> None:
    exam_codes = _get_exam_codes_for_modality(modality=MODALITY_MG)
    _create_report_main(
        template_path=template_path,
        data_count=data_count,
        data_dose=data_dose,
        hospital=hospital,
        modality=MODALITY_MG
    )


def _create_report_file_xa(template_path: Path, data_count: pd.DataFrame, data_dose: pd.DataFrame, hospital: str) -> None:
    exam_codes = _get_exam_codes_for_modality(modality=MODALITY_XA)
    _create_report_main(template_path=template_path, data_count=data_count, data_dose=data_dose, hospital=hospital)